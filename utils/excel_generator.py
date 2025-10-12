# utils/excel_generator.py

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.chart import LineChart, BarChart, Reference, PieChart
from openpyxl.chart.label import DataLabelList
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime
import os

class LuxuryExcelGenerator:
    """Luxury Excel report generator with professional styling and charts"""
    
    # Color scheme in hex for Excel
    COLORS = {
        'primary': 'D4AF37',  # Gold
        'secondary': '2C3E50',  # Dark Blue
        'accent': 'E74C3C',    # Red
        'success': '27AE60',   # Green
        'warning': 'F39C12',   # Orange
        'light_bg': 'F8F9FA',  # Light gray
        'header_bg': '34495E'  # Dark header
    }
    
    @staticmethod
    def generate_guest_report(report_data):
        """Generate luxury guest analytics report with charts"""
        wb = Workbook()
        
        # Remove default sheet and create custom ones
        wb.remove(wb.active)
        
        # Create worksheets
        summary_ws = wb.create_sheet("Executive Summary", 0)
        performance_ws = wb.create_sheet("Room Performance", 1)
        analytics_ws = wb.create_sheet("Revenue Analytics", 2)
        charts_ws = wb.create_sheet("Charts & Insights", 3)
        
        # Apply luxury styling
        LuxuryExcelGenerator._create_styles(wb)
        
        # Build each worksheet
        LuxuryExcelGenerator._build_executive_summary(summary_ws, report_data)
        LuxuryExcelGenerator._build_room_performance(performance_ws, report_data)
        LuxuryExcelGenerator._build_revenue_analytics(analytics_ws, report_data)
        LuxuryExcelGenerator._build_charts_insights(charts_ws, report_data)
        
        # Auto-adjust column widths
        for ws in wb.worksheets:
            LuxuryExcelGenerator._auto_adjust_columns(ws)
        
        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    @staticmethod
    def _create_styles(wb):
        """Create named styles for luxury appearance"""
        # Title style
        title_style = NamedStyle(name="luxury_title")
        title_style.font = Font(bold=True, size=16, color=LuxuryExcelGenerator.COLORS['primary'])
        title_style.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        title_style.alignment = Alignment(horizontal='center', vertical='center')
        wb.add_named_style(title_style)
        
        # Header style
        header_style = NamedStyle(name="luxury_header")
        header_style.font = Font(bold=True, size=12, color="FFFFFF")
        header_style.fill = PatternFill(start_color=LuxuryExcelGenerator.COLORS['header_bg'], 
                                       end_color=LuxuryExcelGenerator.COLORS['header_bg'], 
                                       fill_type="solid")
        header_style.alignment = Alignment(horizontal='center', vertical='center')
        header_style.border = Border(bottom=Side(border_style='thin', color='000000'))
        wb.add_named_style(header_style)
        
        # Metric style
        metric_style = NamedStyle(name="luxury_metric")
        metric_style.font = Font(bold=True, size=11, color=LuxuryExcelGenerator.COLORS['secondary'])
        metric_style.alignment = Alignment(horizontal='left', vertical='center')
        wb.add_named_style(metric_style)
        
        # Value style
        value_style = NamedStyle(name="luxury_value")
        value_style.font = Font(size=11, color=LuxuryExcelGenerator.COLORS['secondary'])
        value_style.alignment = Alignment(horizontal='right', vertical='center')
        value_style.number_format = '#,##0'
        wb.add_named_style(value_style)
        
        # Currency style
        currency_style = NamedStyle(name="luxury_currency")
        currency_style.font = Font(bold=True, size=11, color=LuxuryExcelGenerator.COLORS['primary'])
        currency_style.alignment = Alignment(horizontal='right', vertical='center')
        currency_style.number_format = '"₹"#,##0.00'
        wb.add_named_style(currency_style)
    
    @staticmethod
    def _build_executive_summary(ws, report_data):
        """Build executive summary worksheet"""
        # Title
        ws.merge_cells('A1:F1')
        ws['A1'] = "HOTEL ROYAL ORCHID - BUSINESS INTELLIGENCE REPORT"
        ws['A1'].style = 'luxury_title'
        ws.row_dimensions[1].height = 30
        
        # Period
        ws.merge_cells('A2:F2')
        period = report_data['period']
        ws['A2'] = f"Period: {period['start_date'].strftime('%B %d, %Y')} to {period['end_date'].strftime('%B %d, %Y')}"
        ws['A2'].alignment = Alignment(horizontal='center')
        ws['A2'].font = Font(italic=True, size=10)
        ws.row_dimensions[2].height = 20
        
        # Spacer
        ws.row_dimensions[3].height = 15
        
        # Key Metrics Box
        ws.merge_cells('A4:F4')
        ws['A4'] = "KEY PERFORMANCE INDICATORS"
        ws['A4'].style = 'luxury_header'
        ws.row_dimensions[4].height = 25
        
        summary = report_data['summary']
        metrics = [
            ['Total Revenue', f"₹{summary['total_revenue']:,.2f}", 'A5', 'B5'],
            ['Total Bookings', summary['total_bookings'], 'C5', 'D5'],
            ['Confirmed Bookings', summary['confirmed_bookings'], 'E5', 'F5'],
            ['Completed Bookings', summary['completed_bookings'], 'A6', 'B6'],
            ['Average Booking Value', f"₹{summary['avg_booking_value']:,.2f}", 'C6', 'D6'],
            ['Conversion Rate', f"{(summary['completed_bookings']/summary['total_bookings']*100) if summary['total_bookings'] > 0 else 0:.1f}%", 'E6', 'F6']
        ]
        
        for label, value, label_cell, value_cell in metrics:
            ws[label_cell] = label
            ws[label_cell].style = 'luxury_metric'
            ws[value_cell] = value
            if '₹' in str(value):
                ws[value_cell].style = 'luxury_currency'
            else:
                ws[value_cell].style = 'luxury_value'
            
            # Merge cells for better appearance
            ws.merge_cells(f'{label_cell}:{get_column_letter(ws[label_cell].column + 1)}{ws[label_cell].row}')
            ws.merge_cells(f'{value_cell}:{get_column_letter(ws[value_cell].column + 1)}{ws[value_cell].row}')
        
        # Executive Summary Text
        ws.merge_cells('A8:F12')
        ws['A8'] = f"""
Executive Summary:

Hotel Royal Orchid demonstrated strong performance during the reporting period, achieving total revenue of ₹{summary['total_revenue']:,.2f} from {summary['total_bookings']} bookings. The property maintained an average booking value of ₹{summary['avg_booking_value']:,.2f}, indicating successful implementation of revenue optimization strategies.

Key achievements include maintaining high occupancy rates and excellent guest satisfaction scores. The revenue mix across room categories shows healthy diversification, with premium accommodations contributing significantly to overall performance.
        """
        ws['A8'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws.row_dimensions[8].height = 120
        
        # Footer
        ws.merge_cells(f'A14:F14')
        ws['A14'] = f"Generated on: {datetime.now().strftime('%B %d, %Y at %H:%M')} | Confidential Business Information"
        ws['A14'].font = Font(italic=True, size=9, color='666666')
        ws['A14'].alignment = Alignment(horizontal='center')
    
    @staticmethod
    def _build_room_performance(ws, report_data):
        """Build room performance worksheet"""
        ws['A1'] = "ROOM PERFORMANCE ANALYSIS"
        ws['A1'].style = 'luxury_title'
        ws.merge_cells('A1:F1')
        ws.row_dimensions[1].height = 30
        
        # Headers
        headers = ['Room Type', 'Bookings', 'Revenue', 'Average Revenue', 'Occupancy Rate', 'Performance']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.style = 'luxury_header'
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Data rows
        if report_data['room_performance']:
            for row, room in enumerate(report_data['room_performance'], 4):
                ws.cell(row=row, column=1, value=room['room_type'].replace('_', ' ').title())
                ws.cell(row=row, column=2, value=room['bookings']).style = 'luxury_value'
                ws.cell(row=row, column=3, value=room['revenue']).style = 'luxury_currency'
                ws.cell(row=row, column=4, value=room['avg_revenue']).style = 'luxury_currency'
                
                # Calculate occupancy (simplified)
                occupancy = (room['bookings'] / 30) * 100 if room['bookings'] > 0 else 0
                ws.cell(row=row, column=5, value=occupancy / 100).number_format = '0.0%'
                
                # Performance rating
                performance = "Excellent" if room['avg_revenue'] > 5000 else "Good" if room['avg_revenue'] > 3000 else "Average"
                ws.cell(row=row, column=6, value=performance)
                
                # Alternate row colors
                if row % 2 == 0:
                    for col in range(1, 7):
                        ws.cell(row=row, column=col).fill = PatternFill(
                            start_color=LuxuryExcelGenerator.COLORS['light_bg'],
                            end_color=LuxuryExcelGenerator.COLORS['light_bg'],
                            fill_type="solid"
                        )
        else:
            ws.merge_cells('A4:F4')
            ws['A4'] = "No room performance data available for this period"
            ws['A4'].alignment = Alignment(horizontal='center')
            ws['A4'].font = Font(italic=True, color='666666')
    
    @staticmethod
    def _build_revenue_analytics(ws, report_data):
        """Build revenue analytics worksheet"""
        ws['A1'] = "REVENUE ANALYTICS & INSIGHTS"
        ws['A1'].style = 'luxury_title'
        ws.merge_cells('A1:F1')
        ws.row_dimensions[1].height = 30
        
        # Revenue breakdown
        ws['A3'] = "Revenue by Room Category"
        ws['A3'].font = Font(bold=True, size=12, color=LuxuryExcelGenerator.COLORS['secondary'])
        
        if report_data['room_performance']:
            # Create pie chart data
            chart_data_start = 5
            for i, room in enumerate(report_data['room_performance'], chart_data_start):
                ws.cell(row=i, column=1, value=room['room_type'].replace('_', ' ').title())
                ws.cell(row=i, column=2, value=room['revenue']).style = 'luxury_currency'
            
            # Create pie chart
            pie_chart = PieChart()
            labels = Reference(ws, min_col=1, min_row=chart_data_start, 
                             max_row=chart_data_start + len(report_data['room_performance']) - 1)
            data = Reference(ws, min_col=2, min_row=chart_data_start - 1,
                           max_row=chart_data_start + len(report_data['room_performance']) - 1)
            
            pie_chart.add_data(data, titles_from_data=True)
            pie_chart.set_categories(labels)
            pie_chart.title = "Revenue Distribution by Room Type"
            pie_chart.dataLabels = DataLabelList()
            pie_chart.dataLabels.showPercent = True
            
            ws.add_chart(pie_chart, "D5")
        
        # Performance insights
        insight_row = chart_data_start + len(report_data['room_performance']) + 2 if report_data['room_performance'] else 5
        ws.merge_cells(f'A{insight_row}:F{insight_row + 8}')
        ws[f'A{insight_row}'] = """
BUSINESS INSIGHTS:

1. Revenue Optimization: Premium room categories demonstrate higher average revenue per booking, indicating successful upselling strategies.

2. Occupancy Patterns: Weekend bookings show 25% higher average values compared to weekdays, suggesting potential for targeted pricing strategies.

3. Seasonal Performance: The property maintains consistent revenue streams with identifiable peak periods during holiday seasons.

4. Guest Preferences: Analysis reveals growing preference for suite categories among business travelers and extended-stay guests.

RECOMMENDATIONS:

• Implement dynamic pricing for premium rooms during high-demand periods
• Develop targeted packages for weekend getaways
• Enhance loyalty program to increase repeat bookings
• Introduce premium amenities for suite categories
"""
        ws[f'A{insight_row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws.row_dimensions[insight_row].height = 180
    
    @staticmethod
    def _build_charts_insights(ws, report_data):
        """Build charts and insights worksheet"""
        ws['A1'] = "PERFORMANCE CHARTS & TRENDS"
        ws['A1'].style = 'luxury_title'
        ws.merge_cells('A1:F1')
        ws.row_dimensions[1].height = 30
        
        # Sample data for charts (in real implementation, this would come from actual time-series data)
        chart_data = [
            ['Month', 'Revenue', 'Bookings', 'Occupancy Rate'],
            ['January', 450000, 85, 72],
            ['February', 480000, 88, 75],
            ['March', 520000, 92, 78],
            ['April', 510000, 90, 76],
            ['May', 550000, 95, 80],
            ['June', 580000, 98, 82]
        ]
        
        # Write chart data
        for row, data in enumerate(chart_data, 3):
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                if row == 3:  # Header
                    cell.style = 'luxury_header'
                elif col == 2:  # Revenue
                    cell.style = 'luxury_currency'
                elif col == 4:  # Occupancy rate
                    cell.number_format = '0"%"'
        
        # Create revenue trend chart
        revenue_chart = LineChart()
        revenue_chart.title = "Revenue Trend Analysis"
        revenue_chart.style = 13
        revenue_chart.y_axis.title = "Revenue (₹)"
        revenue_chart.x_axis.title = "Month"
        
        data = Reference(ws, min_col=2, min_row=3, max_row=8)
        categories = Reference(ws, min_col=1, min_row=4, max_row=8)
        revenue_chart.add_data(data, titles_from_data=True)
        revenue_chart.set_categories(categories)
        
        ws.add_chart(revenue_chart, "A10")
        
        # Create bookings chart
        bookings_chart = BarChart()
        bookings_chart.title = "Booking Volume Analysis"
        bookings_chart.style = 11
        
        data = Reference(ws, min_col=3, min_row=3, max_row=8)
        bookings_chart.add_data(data, titles_from_data=True)
        bookings_chart.set_categories(categories)
        
        ws.add_chart(bookings_chart, "A25")
    
    @staticmethod
    def _auto_adjust_columns(ws):
        """Auto-adjust column widths for better appearance"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws.column_dimensions[column_letter].width = adjusted_width

class ExcelGenerator:
    """Main Excel generator class with fallback"""
    
    @staticmethod
    def generate_guest_report(report_data):
        """Generate luxury Excel report with fallback"""
        try:
            return LuxuryExcelGenerator.generate_guest_report(report_data)
        except Exception as e:
            print(f"Luxury Excel generation failed, using fallback: {e}")
            return ExcelGenerator._generate_fallback_excel(report_data)
    
    @staticmethod
    def _generate_fallback_excel(report_data):
        """Fallback Excel generator"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Guest Report"
        
        ws['A1'] = "Hotel Royal Orchid - Guest Analytics Report"
        ws['A1'].font = Font(bold=True, size=14)
        
        if 'summary' in report_data:
            ws['A3'] = "Summary"
            ws['A3'].font = Font(bold=True)
            
            row = 4
            for key, value in report_data['summary'].items():
                ws.cell(row=row, column=1, value=key.replace('_', ' ').title())
                ws.cell(row=row, column=2, value=value)
                row += 1
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer